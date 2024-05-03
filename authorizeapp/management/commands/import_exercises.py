from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from authorizeapp.models import ExerciseList

class Command(BaseCommand):
    help = 'Populate exercises from Excel file'

    def handle(self, *args, **kwargs):
        wb = load_workbook('Corrected_Gym_Exercises.xlsx')
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            exercise_name, muscle_gp, Equipment, *_ = row

            if(exercise_name):

                ExerciseList.objects.create(
                    name=exercise_name,
                    muscle_group=muscle_gp,
                    equipment=Equipment,
                )
            else: return